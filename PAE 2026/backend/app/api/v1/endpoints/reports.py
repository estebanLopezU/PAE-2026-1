from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from collections import Counter, defaultdict
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from ....database import get_db
from ....models.entity import Entity
from ....models.service import Service
from ....models.maturity import MaturityAssessment
from ....models.sector import Sector

router = APIRouter()


def _apply_corporate_header(ws, report_title: str, subtitle: str):
    ws.merge_cells("A1:F1")
    ws["A1"] = "X-Road Colombia - Reporte Ejecutivo"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="EEF3F8")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = report_title
    ws["A2"].font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = subtitle
    ws["A3"].font = Font(name="Calibri", size=10, color="4F4F4F")
    ws["A3"].alignment = Alignment(horizontal="center")


def _apply_data_sheet_style(ws, headers: list[str], rows_count: int):
    header_row = 4
    data_start_row = 5
    max_col = len(headers)
    max_row = data_start_row + rows_count - 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="Detalle de Datos")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1F4E78")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, color="666666")

    # Header style
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(name="Calibri", bold=True, color="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data cells style
    for row_idx in range(data_start_row, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # Freeze panes and filters
    ws.freeze_panes = f"A{data_start_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max(data_start_row, max_row)}"

    # Excel table style
    if rows_count > 0:
        table_ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=f"Table_{ws.title.replace(' ', '_')}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Auto width
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(14, max_length + 2), 42)


def _normalize_xroad_status(status: str) -> str:
    value = (status or "not_connected").strip().lower()
    mapping = {
        "connected": "Conectada",
        "active": "Conectada",
        "conectado": "Conectada",
        "conectada": "Conectada",
        "pending": "En Progreso",
        "in_progress": "En Progreso",
        "progreso": "En Progreso",
        "not_connected": "No Conectada",
        "inactive": "No Conectada",
        "desconectada": "No Conectada",
    }
    return mapping.get(value, value.title())


def _connection_score(status: str) -> int:
    normalized = _normalize_xroad_status(status)
    if normalized == "Conectada":
        return 100
    if normalized == "En Progreso":
        return 60
    return 20


def _safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value if value is not None else default)
    except Exception:
        return default


def _score_classification(score: float) -> str:
    if score >= 85:
        return "A - Excelente"
    if score >= 70:
        return "B - Bueno"
    if score >= 50:
        return "C - Atención"
    return "D - Crítico"


def _priority_level(score: float) -> str:
    if score < 50:
        return "Alta"
    if score < 70:
        return "Media"
    return "Baja"


def _status_fill(value: str) -> PatternFill:
    v = (value or "").lower()
    if "excelente" in v or "conectada" in v or "baja" in v:
        return PatternFill(fill_type="solid", fgColor="EAF7EA")
    if "bueno" in v or "progreso" in v or "media" in v:
        return PatternFill(fill_type="solid", fgColor="FFF7DF")
    if "atención" in v or "crítico" in v or "alta" in v or "no conectada" in v:
        return PatternFill(fill_type="solid", fgColor="FDEBEC")
    return PatternFill(fill_type="solid", fgColor="FFFFFF")


def _build_excel_response(
    sheet_name: str,
    headers: list[str],
    rows: list[list],
    filename: str,
    report_title: str,
    summary_items: list[tuple[str, str]],
    extra_sheets: list[dict] | None = None,
):
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Resumen Ejecutivo"

    subtitle = "Documento empresarial para análisis y toma de decisiones"
    _apply_corporate_header(summary_ws, report_title, subtitle)

    start_row = 5
    key_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    value_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, (label, value) in enumerate(summary_items, start=start_row):
        key_cell = summary_ws.cell(row=idx, column=1, value=label)
        value_cell = summary_ws.cell(row=idx, column=2, value=value)

        key_cell.font = Font(name="Calibri", bold=True, color="1F4E78")
        key_cell.fill = key_fill
        key_cell.border = border
        key_cell.alignment = Alignment(vertical="top", wrap_text=True)
        value_cell.font = Font(name="Calibri", size=10)
        value_cell.fill = value_fill
        value_cell.border = border
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary_ws.column_dimensions["A"].width = 36
    summary_ws.column_dimensions["B"].width = 44

    worksheet = workbook.create_sheet(title=sheet_name)

    # Header/data starting rows for data sheet
    worksheet.append([])  # row 1
    worksheet.append([])  # row 2
    worksheet.append([])  # row 3
    worksheet.append(headers)  # row 4
    for row in rows:
        worksheet.append(row)

    _apply_data_sheet_style(worksheet, headers, len(rows))

    # Hojas adicionales (análisis ejecutivo/sectorial/plan de acción)
    for sheet in extra_sheets or []:
        ws = workbook.create_sheet(title=sheet["title"])
        local_headers = sheet.get("headers", [])
        local_rows = sheet.get("rows", [])
        colored_columns = sheet.get("colored_columns", [])

        ws.append([])
        ws.append([])
        ws.append([])
        ws.append(local_headers)
        for row in local_rows:
            ws.append(row)

        _apply_data_sheet_style(ws, local_headers, len(local_rows))

        if local_rows and colored_columns:
            for row_idx in range(5, 5 + len(local_rows)):
                for col_idx in colored_columns:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = _status_fill(str(cell.value))
                    cell.font = Font(name="Calibri", size=10, bold=True)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/entities/csv")
def download_entities_csv(db: Session = Depends(get_db)):
    """Download entities report as CSV"""
    entities = db.query(
        Entity.name,
        Entity.acronym,
        Entity.nit,
        Entity.department,
        Entity.xroad_status,
        Sector.name.label('sector_name')
    ).join(Sector, Entity.sector_id == Sector.id, isouter=True).filter(
        Entity.is_active == True
    ).all()
    
    def normalize_status(status):
        value = (status or "not_connected").strip().lower()
        if value in ["connected", "active", "conectado", "conectada"]:
            return "Conectada"
        if value in ["pending", "in_progress", "progreso"]:
            return "En Proceso"
        return "Sin Conexion"
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Encabezado del reporte
    writer.writerow(['X-ROAD COLOMBIA - REPORTE DE ENTIDADES'])
    writer.writerow([f'Fecha de generacion: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([])
    
    # Resumen
    connected = [e for e in entities if normalize_status(e.xroad_status) == "Conectada"]
    pending = [e for e in entities if normalize_status(e.xroad_status) == "En Proceso"]
    not_connected = [e for e in entities if normalize_status(e.xroad_status) == "Sin Conexion"]
    
    writer.writerow(['RESUMEN'])
    writer.writerow(['Total entidades activas', len(entities)])
    writer.writerow(['Entidades conectadas', len(connected)])
    writer.writerow(['Entidades en proceso', len(pending)])
    writer.writerow(['Entidades sin conexion', len(not_connected)])
    writer.writerow([])
    
    # Separador
    writer.writerow([',' + '='*80])
    writer.writerow([])
    
    # ENTIDADES CONECTADAS
    writer.writerow(['ENTIDADES CONECTADAS A X-ROAD'])
    writer.writerow(['Nombre', 'NIT', 'Ciudad/Departamento', 'Sector', 'Estado'])
    for entity in sorted(connected, key=lambda x: x.name.lower()):
        writer.writerow([
            entity.name,
            entity.nit or 'N/A',
            entity.department or 'N/A',
            entity.sector_name or 'N/A',
            'Conectada'
        ])
    writer.writerow([])
    writer.writerow([',' + '='*80])
    writer.writerow([])
    
    # ENTIDADES PENDIENTES
    writer.writerow(['ENTIDADES EN ESTADO PENDIENTE'])
    writer.writerow(['Nombre', 'NIT', 'Ciudad/Departamento', 'Sector', 'Estado'])
    for entity in sorted(pending, key=lambda x: x.name.lower()):
        writer.writerow([
            entity.name,
            entity.nit or 'N/A',
            entity.department or 'N/A',
            entity.sector_name or 'N/A',
            'En Proceso'
        ])
    writer.writerow([])
    writer.writerow([',' + '='*80])
    writer.writerow([])
    
    # ENTIDADES NO CONECTADAS
    writer.writerow(['ENTIDADES NO CONECTADAS'])
    writer.writerow(['Nombre', 'NIT', 'Ciudad/Departamento', 'Sector', 'Estado'])
    for entity in sorted(not_connected, key=lambda x: x.name.lower()):
        writer.writerow([
            entity.name,
            entity.nit or 'N/A',
            entity.department or 'N/A',
            entity.sector_name or 'N/A',
            'Sin Conexion'
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=reporte_entidades.csv"}
    )


@router.get("/entities/xlsx")
def download_entities_xlsx(db: Session = Depends(get_db)):
    """Download entities report as XLSX with full entity lists"""
    entities = db.query(
        Entity.id,
        Entity.name,
        Entity.acronym,
        Entity.nit,
        Entity.department,
        Entity.xroad_status,
        Sector.name.label('sector_name')
    ).join(Sector, Entity.sector_id == Sector.id, isouter=True).filter(
        Entity.is_active == True
    ).all()

    headers = ['ID', 'Nombre', 'Acrónimo', 'NIT', 'Departamento', 'Estado X-Road', 'Sector']
    rows = [
        [
            entity.id,
            entity.name,
            entity.acronym,
            entity.nit,
            entity.department,
            _normalize_xroad_status(entity.xroad_status),
            entity.sector_name,
        ]
        for entity in entities
    ]

    connected_rows = [row for row in rows if row[5] == "Conectada"]
    pending_rows = [row for row in rows if row[5] == "En Progreso"]
    not_connected_rows = [row for row in rows if row[5] == "No Conectada"]
    total = len(rows) if rows else 1

    connectivity_rate = round((len(connected_rows) / total) * 100, 2) if total else 0

    # Detalle completo con ID, NIT y Ciudad
    connected_detail = []
    for r in sorted(connected_rows, key=lambda x: str(x[1]).lower()):
        connected_detail.append([
            r[1],  # Nombre
            r[3] or "N/A",  # NIT
            r[4] or "N/A",  # Ciudad/Departamento
            r[2] or "N/A",  # Sigla
            r[6] or "N/A",  # Sector
            "CONECTADA"     # Estado
        ])

    pending_detail = []
    for r in sorted(pending_rows, key=lambda x: str(x[1]).lower()):
        pending_detail.append([
            r[1],  # Nombre
            r[3] or "N/A",  # NIT
            r[4] or "N/A",  # Ciudad/Departamento
            r[2] or "N/A",  # Sigla
            r[6] or "N/A",  # Sector
            "PENDIENTE"     # Estado
        ])

    not_connected_detail = []
    for r in sorted(not_connected_rows, key=lambda x: str(x[1]).lower()):
        not_connected_detail.append([
            r[1],  # Nombre
            r[3] or "N/A",  # NIT
            r[4] or "N/A",  # Ciudad/Departamento
            r[2] or "N/A",  # Sigla
            r[6] or "N/A",  # Sector
            "SIN CONEXION"   # Estado
        ])

    # Matriz completa con numeros
    matrix_rows = []
    status_order = {"Conectada": 0, "En Progreso": 1, "No Conectada": 2}
    sorted_all = sorted(rows, key=lambda r: (status_order.get(r[5], 99), str(r[1]).lower()))
    for idx, r in enumerate(sorted_all, start=1):
        matrix_rows.append([
            idx,
            r[1],  # Nombre
            r[3] or "N/A",  # NIT
            r[4] or "N/A",  # Ciudad
            r[2] or "N/A",  # Sigla
            r[6] or "N/A",  # Sector
            r[5],  # Estado
        ])

    return _build_excel_response(
        sheet_name="Resumen",
        headers=headers,
        rows=rows,
        filename="reporte_entidades.xlsx",
        report_title="Reporte Corporativo de Entidades",
        summary_items=[
            ("Fecha de generación", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Total de entidades activas", str(len(rows))),
            ("Entidades conectadas a X-Road", str(len(connected_rows))),
            ("Entidades en estado pendiente", str(len(pending_rows))),
            ("Entidades no conectadas", str(len(not_connected_rows))),
            ("Tasa de conectividad", f"{connectivity_rate}%"),
            ("Formato", "Excel (.xlsx) con listado completo por estado")
        ],
        extra_sheets=[
            {
                "title": "CONECTADAS",
                "headers": ["Nombre Entidad", "Numero Identificacion (NIT)", "Ciudad Departamento", "Sigla", "Sector", "Estado"],
                "rows": connected_detail,
                "colored_columns": [6],
            },
            {
                "title": "PENDIENTES",
                "headers": ["Nombre Entidad", "Numero Identificacion (NIT)", "Ciudad Departamento", "Sigla", "Sector", "Estado"],
                "rows": pending_detail,
                "colored_columns": [6],
            },
            {
                "title": "NO CONECTADAS",
                "headers": ["Nombre Entidad", "Numero Identificacion (NIT)", "Ciudad Departamento", "Sigla", "Sector", "Estado"],
                "rows": not_connected_detail,
                "colored_columns": [6],
            },
            {
                "title": "MATRIZ COMPLETA",
                "headers": ["#", "Nombre Entidad", "NIT", "Ciudad/Departamento", "Sigla", "Sector", "Estado X-Road"],
                "rows": matrix_rows,
                "colored_columns": [7],
            },
        ],
    )


@router.get("/services/csv")
def download_services_csv(db: Session = Depends(get_db)):
    """Download services report as CSV"""
    services = db.query(
        Service.name,
        Service.code,
        Service.description,
        Service.protocol,
        Service.category,
        Service.status,
        Entity.name.label('entity_name'),
        Sector.name.label('sector_name')
    ).join(Entity, Service.entity_id == Entity.id
    ).join(Sector, Entity.sector_id == Sector.id, isouter=True
    ).filter(Service.status == "active").all()
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Encabezado
    writer.writerow(['X-ROAD COLOMBIA - REPORTE DE SERVICIOS'])
    writer.writerow([f'Fecha de generacion: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([])
    
    # Resumen
    protocols = set(s.protocol for s in services if s.protocol)
    categories = set(s.category for s in services if s.category)
    
    writer.writerow(['RESUMEN'])
    writer.writerow(['Total servicios activos', len(services)])
    writer.writerow(['Total protocolos', len(protocols)])
    writer.writerow(['Total categorias', len(categories)])
    writer.writerow([])
    
    writer.writerow([',' + '='*80])
    writer.writerow([])
    
    # Detalle
    writer.writerow(['SERVICIOS DE INTEROPERABILIDAD'])
    writer.writerow(['Nombre Servicio', 'Codigo', 'Entidad', 'Sector', 'Protocolo', 'Categoria'])
    for service in sorted(services, key=lambda x: (x.entity_name or '', x.name)):
        writer.writerow([
            service.name,
            service.code or 'N/A',
            service.entity_name or 'N/A',
            service.sector_name or 'N/A',
            service.protocol or 'N/A',
            service.category or 'N/A'
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=reporte_servicios.csv"}
    )


@router.get("/services/xlsx")
def download_services_xlsx(db: Session = Depends(get_db)):
    """Download services report as XLSX"""
    services = db.query(
        Service.name,
        Service.code,
        Service.description,
        Service.protocol,
        Service.category,
        Service.status,
        Entity.name.label('entity_name')
    ).join(Entity, Service.entity_id == Entity.id).filter(
        Service.status == "active"
    ).all()

    headers = ['Nombre Servicio', 'Código', 'Descripción', 'Protocolo', 'Categoría', 'Estado', 'Entidad']
    rows = [
        [
            service.name,
            service.code,
            service.description,
            service.protocol,
            service.category,
            service.status,
            service.entity_name,
        ]
        for service in services
    ]

    return _build_excel_response(
        sheet_name="Servicios",
        headers=headers,
        rows=rows,
        filename="reporte_servicios.xlsx",
        report_title="Reporte Corporativo de Servicios",
        summary_items=[
            ("Fecha de generación", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Total de servicios activos", str(len(rows))),
            ("Protocolos identificados", str(len(set(row[3] for row in rows if row[3])))),
            ("Categorías identificadas", str(len(set(row[4] for row in rows if row[4])))),
            ("Formato", "Excel (.xlsx) empresarial con resumen y detalle"),
        ],
    )


@router.get("/maturity/csv")
def download_maturity_csv(db: Session = Depends(get_db)):
    """Download maturity assessments report as CSV"""
    # Get all assessments with entity information
    assessments = db.query(
        MaturityAssessment.id,
        Entity.name.label('entity_name'),
        MaturityAssessment.overall_level,
        MaturityAssessment.overall_score,
        MaturityAssessment.legal_domain_score,
        MaturityAssessment.organizational_domain_score,
        MaturityAssessment.semantic_domain_score,
        MaturityAssessment.technical_domain_score,
        MaturityAssessment.has_api_documentation,
        MaturityAssessment.uses_standard_protocols,
        MaturityAssessment.has_data_quality,
        MaturityAssessment.has_security_standards,
        MaturityAssessment.has_interoperability_policy,
        MaturityAssessment.has_trained_personnel,
        MaturityAssessment.assessor_name,
        MaturityAssessment.assessment_date
    ).join(Entity, MaturityAssessment.entity_id == Entity.id).all()
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'ID', 'Entidad', 'Nivel General', 'Puntuación General',
        'Dominio Legal', 'Dominio Organizacional', 'Dominio Semántico', 'Dominio Técnico',
        'Documentación APIs', 'Protocolos Estándar', 'Calidad Datos',
        'Seguridad', 'Política Interoperabilidad', 'Personal Capacitado',
        'Evaluador', 'Fecha Evaluación'
    ])
    
    # Write data
    for assessment in assessments:
        writer.writerow([
            assessment.id,
            assessment.entity_name,
            assessment.overall_level,
            assessment.overall_score,
            assessment.legal_domain_score,
            assessment.organizational_domain_score,
            assessment.semantic_domain_score,
            assessment.technical_domain_score,
            assessment.has_api_documentation,
            assessment.uses_standard_protocols,
            assessment.has_data_quality,
            assessment.has_security_standards,
            assessment.has_interoperability_policy,
            assessment.has_trained_personnel,
            assessment.assessor_name,
            assessment.assessment_date
        ])
    
    # Prepare response
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=reporte_madurez.csv"}
    )


@router.get("/maturity/xlsx")
def download_maturity_xlsx(db: Session = Depends(get_db)):
    """Download maturity assessments report as XLSX"""
    assessments = db.query(
        MaturityAssessment.id,
        Entity.name.label('entity_name'),
        MaturityAssessment.overall_level,
        MaturityAssessment.overall_score,
        MaturityAssessment.legal_domain_score,
        MaturityAssessment.organizational_domain_score,
        MaturityAssessment.semantic_domain_score,
        MaturityAssessment.technical_domain_score,
        MaturityAssessment.has_api_documentation,
        MaturityAssessment.uses_standard_protocols,
        MaturityAssessment.has_data_quality,
        MaturityAssessment.has_security_standards,
        MaturityAssessment.has_interoperability_policy,
        MaturityAssessment.has_trained_personnel,
        MaturityAssessment.assessor_name,
        MaturityAssessment.assessment_date
    ).join(Entity, MaturityAssessment.entity_id == Entity.id).all()

    headers = [
        'ID', 'Entidad', 'Nivel General', 'Puntuación General',
        'Dominio Legal', 'Dominio Organizacional', 'Dominio Semántico', 'Dominio Técnico',
        'Documentación APIs', 'Protocolos Estándar', 'Calidad Datos',
        'Seguridad', 'Política Interoperabilidad', 'Personal Capacitado',
        'Evaluador', 'Fecha Evaluación'
    ]
    rows = [
        [
            assessment.id,
            assessment.entity_name,
            assessment.overall_level,
            assessment.overall_score,
            assessment.legal_domain_score,
            assessment.organizational_domain_score,
            assessment.semantic_domain_score,
            assessment.technical_domain_score,
            assessment.has_api_documentation,
            assessment.uses_standard_protocols,
            assessment.has_data_quality,
            assessment.has_security_standards,
            assessment.has_interoperability_policy,
            assessment.has_trained_personnel,
            assessment.assessor_name,
            assessment.assessment_date.replace(tzinfo=None) if assessment.assessment_date else None,
        ]
        for assessment in assessments
    ]

    # =====================
    # Análisis profesional por entidad (última evaluación)
    # =====================
    entities = db.query(Entity).filter(Entity.is_active == True).all()

    entity_rows = []
    action_rows = []
    connection_counter = Counter()
    sector_acc = defaultdict(lambda: {"count": 0, "connected": 0, "score_sum": 0.0})

    for entity in entities:
        latest = db.query(MaturityAssessment).filter(
            MaturityAssessment.entity_id == entity.id
        ).order_by(MaturityAssessment.assessment_date.desc()).first()

        maturity_score = _safe_float(latest.overall_score if latest else 0)
        conn_status = _normalize_xroad_status(entity.xroad_status)
        conn_score = _connection_score(entity.xroad_status)

        # Índice compuesto: 60% madurez, 25% conexión, 15% documentación/API
        doc_component = _safe_float((latest.has_api_documentation * 25) if latest else 0)
        composite_score = round(maturity_score * 0.60 + conn_score * 0.25 + doc_component * 0.15, 2)

        classification = _score_classification(composite_score)
        priority = _priority_level(composite_score)
        sector_name = entity.sector.name if entity.sector else "Sin Sector"

        entity_rows.append([
            entity.name,
            sector_name,
            conn_status,
            round(maturity_score, 2),
            conn_score,
            composite_score,
            classification,
            priority,
        ])

        # Plan de acción 30/60/90
        if priority == "Alta":
            action, owner, timeframe, impact = (
                "Implementar estándares API + plan de conexión X-Road",
                "CIO / Arquitectura TI",
                "30 días",
                "Alto",
            )
        elif priority == "Media":
            action, owner, timeframe, impact = (
                "Mejorar calidad semántica y documentación OpenAPI",
                "Líder de Datos / Integraciones",
                "60 días",
                "Medio-Alto",
            )
        else:
            action, owner, timeframe, impact = (
                "Optimización continua y monitoreo de interoperabilidad",
                "Equipo de Interoperabilidad",
                "90 días",
                "Medio",
            )

        action_rows.append([
            entity.name,
            sector_name,
            priority,
            action,
            owner,
            timeframe,
            impact,
        ])

        connection_counter[conn_status] += 1
        sector_acc[sector_name]["count"] += 1
        sector_acc[sector_name]["score_sum"] += composite_score
        if conn_status == "Conectada":
            sector_acc[sector_name]["connected"] += 1

    total_entities = len(entity_rows) if entity_rows else 1
    connection_rows = []
    for status_name in ["Conectada", "En Progreso", "No Conectada"]:
        count = connection_counter.get(status_name, 0)
        pct = round(count / total_entities * 100, 2)
        connection_rows.append([status_name, count, f"{pct}%"])

    sector_rows = []
    for sector_name, data in sorted(sector_acc.items(), key=lambda x: x[0]):
        avg_score = round(data["score_sum"] / data["count"], 2) if data["count"] else 0
        coverage = round(data["connected"] / data["count"] * 100, 2) if data["count"] else 0
        sector_rows.append([sector_name, data["count"], data["connected"], f"{coverage}%", avg_score])

    technical_annex_rows = [
        ["Índice Compuesto", "0-100", "0.60*Madurez + 0.25*Conexión + 0.15*Documentación"],
        ["Conexión", "20/60/100", "No conectada=20, En progreso=60, Conectada=100"],
        ["Clasificación A", ">=85", "Excelente preparación institucional"],
        ["Clasificación B", "70-84.99", "Buen nivel, requiere ajustes puntuales"],
        ["Clasificación C", "50-69.99", "Atención: brechas relevantes"],
        ["Clasificación D", "<50", "Crítico: intervención prioritaria"],
        ["Prioridad Alta", "<50", "Intervenir en 30 días"],
        ["Prioridad Media", "50-69.99", "Intervenir en 60 días"],
        ["Prioridad Baja", ">=70", "Plan de optimización a 90 días"],
    ]

    avg_composite = round(sum(r[5] for r in entity_rows) / len(entity_rows), 2) if entity_rows else 0.0

    return _build_excel_response(
        sheet_name="Madurez",
        headers=headers,
        rows=rows,
        filename="reporte_madurez.xlsx",
        report_title="Reporte Corporativo de Madurez",
        summary_items=[
            ("Fecha de generación", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Total de evaluaciones", str(len(rows))),
            ("Total de entidades activas", str(len(entity_rows))),
            (
                "Promedio de puntuación general",
                f"{(sum(float(row[3] or 0) for row in rows) / len(rows)):.2f}" if rows else "0.00",
            ),
            ("Índice compuesto promedio", f"{avg_composite:.2f}"),
            ("Cobertura de conexión X-Road", f"{round(connection_counter.get('Conectada', 0) / total_entities * 100, 2)}%"),
            ("Nivel máximo detectado", str(max((int(row[2]) for row in rows), default=0))),
            ("Formato", "Excel (.xlsx) empresarial con resumen y detalle"),
        ],
        extra_sheets=[
            {
                "title": "Indice Entidades",
                "headers": [
                    "Entidad",
                    "Sector",
                    "Estado Conexión",
                    "Puntaje Madurez",
                    "Score Conexión",
                    "Índice Compuesto",
                    "Clasificación",
                    "Prioridad",
                ],
                "rows": entity_rows,
                "colored_columns": [3, 7, 8],
            },
            {
                "title": "Conectividad",
                "headers": ["Estado", "Total Entidades", "Porcentaje"],
                "rows": connection_rows,
                "colored_columns": [1],
            },
            {
                "title": "Analisis Sectorial",
                "headers": ["Sector", "Total", "Conectadas", "Cobertura", "Índice Promedio"],
                "rows": sector_rows,
            },
            {
                "title": "Plan 30-60-90",
                "headers": [
                    "Entidad",
                    "Sector",
                    "Prioridad",
                    "Acción Recomendada",
                    "Responsable",
                    "Plazo",
                    "Impacto Esperado",
                ],
                "rows": action_rows,
                "colored_columns": [3],
            },
            {
                "title": "Anexo Tecnico",
                "headers": ["Métrica", "Rango", "Descripción"],
                "rows": technical_annex_rows,
            },
        ],
    )